import openpyxl

# Giving the location of the file
path = 'IllustrativeLifeTable.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

Type = input('Life or Annuity?').strip()
Age = int(input('How old are you?').strip())
Pay = float(input('What is your payout?').strip())
Row = int(Age - 19)

# The spreadsheet does not work for <20 or >100
if Age < 20 or Age > 100:
    print('Sorry, you are out of the age range.')

# Selecting the right column
elif Type == 'Life' or Type == 'life':
    benefit = ws.cell(row=Row, column=5)
    px_1 = 1 - ws.cell(row=Row, column=3).value
    px_2 = 1 - ws.cell(row=Row+1, column=3).value
    px_3 = 1 - ws.cell(row=Row+2, column=3).value
    px_4 = 1 - ws.cell(row=Row+3, column=3).value
    premium = Pay * benefit.value / (1 + px_1/1.05 + px_2/1.05**2 + px_3/1.05**3 + px_4/1.05**4)
    print('Your annual premium is: ', round(premium, 2))

elif Type == 'Annuity' or Type == 'annuity':
    benefit = ws.cell(row=Row, column=4)
    px_1 = 1 - ws.cell(row=Row, column=3).value
    px_2 = 1 - ws.cell(row=Row + 1, column=3).value
    px_3 = 1 - ws.cell(row=Row + 2, column=3).value
    px_4 = 1 - ws.cell(row=Row + 3, column=3).value
    premium = Pay * benefit.value / (1 + px_1 / 1.05 + px_2 / 1.05 ** 2 + px_3 / 1.05 ** 3 + px_4 / 1.05 ** 4)
    print('Your annual premium is: ', round(premium, 2))

print('Thank you for using Find My Premium!')