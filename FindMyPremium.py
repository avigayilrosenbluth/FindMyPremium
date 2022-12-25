import openpyxl

# Giving the location of the file
path = 'IllustrativeLifeTable.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb.active

Type = input('Life or Annuity?').strip()
Age = int(input('How old are you?').strip())
Pay = float(input('What is your payout?').strip())
Row = int(Age - 19)

# The spreadsheet does not work for <20 or >100
if Age < 20 or Age > 100:
    print('Sorry, you are out of the age range.')

# Selecting the right column
elif Type == 'Life' or Type == 'life':
    cell = ws.cell(row=Row, column=5)
    print('Your premium is: ' + str(Pay * cell.value))
elif Type == 'Annuity' or Type == 'annuity':
    cell = ws.cell(row=Row, column=4)
    print('Your premium is: ' + str(Pay * cell.value))
print('Thank you for using Find My Premium!')